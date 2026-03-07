from __future__ import annotations

from collections.abc import Iterable, Iterator
from pathlib import Path

from ._normalize import normalize_whitespace_line
from .base import ConversionResult


def _iter_compacted_rows(
    rows: Iterable[tuple[object, ...]], *, max_blank_lines: int = 1
) -> Iterator[str]:
    pending_blank_lines = 0
    seen_content = False

    for row in rows:
        cells = [normalize_whitespace_line("" if value is None else str(value)) for value in row]
        while cells and not cells[-1].strip():
            cells.pop()

        line = "\t".join(cells)
        if not line.strip():
            if seen_content:
                pending_blank_lines = min(pending_blank_lines + 1, max_blank_lines)
            continue

        seen_content = True
        for _ in range(pending_blank_lines):
            yield ""
        pending_blank_lines = 0
        yield line


class XlsxFallbackConverter:
    def can_convert(self, ext: str) -> bool:
        try:
            import openpyxl  # noqa: F401

            return ext.lower() == ".xlsx"
        except ImportError:
            return False

    def convert(self, path: Path, encoding: str = "utf-8") -> ConversionResult:
        import openpyxl

        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        parts: list[str] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            parts.append(f"## Sheet: {sheet_name}")
            body = "\n".join(_iter_compacted_rows(ws.iter_rows(values_only=True)))
            if body:
                parts.append(body)
        return ConversionResult(
            content="\n\n".join(parts),
            converter_name="openpyxl",
            original_mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
